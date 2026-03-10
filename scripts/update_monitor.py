import argparse
import csv
import io
import re
import shutil
import statistics
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.parse import quote_plus

import openpyxl
import requests
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from requests.exceptions import RequestException

REQUEST_TIMEOUT = 40
RETRIES = 2
RETRY_WAIT_SECONDS = 2

TEMPLATE_NAME = "private_credit_monitor_template.xlsx"

ROW_IDS = [
    "EW-01", "EW-02", "EW-03", "EW-04", "EW-05",
    "TR-01", "TR-02", "TR-03", "TR-04", "TR-05", "TR-06",
    "SY-01", "SY-02", "SY-03", "SY-04", "SY-05",
]

NEWS_RULES = {
    "EW-01": {
        "query": '"private credit" gate OR "redemption suspension" OR "withdrawal suspension" OR "redemption limit"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "限赎 / Gate",
        "lookback_days": 45,
        "yellow_min_articles": 1,
        "yellow_min_sources": 1,
        "red_min_articles": 3,
        "red_min_sources": 3,
        "must_have_any": [
            "gate", "redemption suspension", "withdrawal suspension",
            "redemption limit", "limits withdrawals", "limited withdrawals",
            "limits redemptions", "redemptions at"
        ],
    },
    "EW-05": {
        "query": '"private credit" fraud OR "valuation dispute" OR default OR restructuring OR "payment suspension"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "坏消息簇发",
        "lookback_days": 21,
        "yellow_min_articles": 2,
        "yellow_min_sources": 2,
        "red_min_articles": 5,
        "red_min_sources": 4,
        "must_have_any": [
            "fraud", "valuation dispute", "default", "restructuring",
            "payment suspension", "writes down", "writedown", "write-down",
            "credit jitters", "distress"
        ],
    },
    "TR-01": {
        "query": '"private credit" default OR restructuring OR "amend and extend" OR "debt exchange"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "违约 / 重组 / 展期",
        "lookback_days": 30,
        "yellow_min_articles": 2,
        "yellow_min_sources": 2,
        "red_min_articles": 4,
        "red_min_sources": 3,
        "must_have_any": [
            "default", "restructuring", "amend and extend",
            "extend maturities", "debt exchange", "debt restructuring",
            "missed payment"
        ],
    },
    "TR-02": {
        "query": '"private credit" PIK OR "paid in kind" OR "payment in kind"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "PIK 增多",
        "lookback_days": 45,
        "yellow_min_articles": 1,
        "yellow_min_sources": 1,
        "red_min_articles": 3,
        "red_min_sources": 3,
        "must_have_any": ["pik", "paid in kind", "payment in kind"],
    },
    "TR-04": {
        "query": '"private credit" fundraising slowdown OR "fundraising slows" OR "difficult to raise" OR "fund close delayed"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "募资放缓",
        "lookback_days": 45,
        "yellow_min_articles": 2,
        "yellow_min_sources": 2,
        "red_min_articles": 4,
        "red_min_sources": 3,
        "must_have_any": [
            "fundraising slowdown", "fundraising slows", "difficult to raise",
            "fund close delayed", "slower fundraising", "harder to raise",
            "fundraising environment"
        ],
    },
    "TR-05": {
        "query": '"private credit" tighter terms OR covenant OR "spread wider" OR "deal pulled" OR "financing delayed"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "条款收紧",
        "lookback_days": 30,
        "yellow_min_articles": 2,
        "yellow_min_sources": 2,
        "red_min_articles": 4,
        "red_min_sources": 3,
        "must_have_any": [
            "tighter terms", "covenant", "spread wider", "deal pulled",
            "financing delayed", "tougher terms", "higher spreads",
            "stricter terms"
        ],
    },
    "SY-03": {
        "query": '"private credit" pension reduce allocation OR insurer reduce allocation OR "trim exposure" OR "cut allocation"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "长钱减配",
        "lookback_days": 60,
        "yellow_min_articles": 1,
        "yellow_min_sources": 1,
        "red_min_articles": 3,
        "red_min_sources": 3,
        "must_have_any": [
            "reduce allocation", "trim exposure", "cut allocation",
            "lower allocation", "reduce exposure"
        ],
    },
    "SY-05": {
        "query": '"small business" refinancing failure OR bankruptcy OR layoffs OR capex cuts OR "unable to refinance"',
        "automation": "新闻聚类",
        "source": "Google News RSS / 主流媒体",
        "label": "中小企业再融资失败",
        "lookback_days": 30,
        "yellow_min_articles": 3,
        "yellow_min_sources": 2,
        "red_min_articles": 6,
        "red_min_sources": 4,
        "must_have_any": [
            "unable to refinance", "refinancing failure", "bankruptcy",
            "layoffs", "capex cuts", "cannot refinance",
            "struggles to refinance"
        ],
    },
}

MANAGER_STOCKS = ["BLK", "BX", "APO", "ARES", "OWL", "KKR"]
ALL_STOOQ = [
    "SPY", "XLF", "KBE", "BLK", "BX", "APO", "ARES", "OWL", "KKR",
    "BIZD", "HYG", "BKLN", "JBBB", "LQD", "JNK",
]
FRED_SERIES = {
    "HY_OAS": "BAMLH0A0HYM2",
    "IG_OAS": "BAMLC0A0CM",
    "SOFR": "SOFR",
    "ANFCI": "ANFCI",
}

LEVEL_STYLE = {
    "绿": {"label": "稳定", "fill": "E2F0D9", "font": "2F6B2F"},
    "浅黄": {"label": "预警升温", "fill": "FFF2CC", "font": "9C6500"},
    "黄": {"label": "压力上升", "fill": "F4E1A1", "font": "7F6000"},
    "橙": {"label": "系统风险临近", "fill": "FCE4D6", "font": "C55A11"},
    "红": {"label": "危机确认", "fill": "F4CCCC", "font": "9C0006"},
}


def fetch_text(url: str) -> str:
    headers = {"User-Agent": "Mozilla/5.0"}
    last_err = None
    for attempt in range(RETRIES + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            return resp.text
        except RequestException as exc:
            last_err = exc
            if attempt < RETRIES:
                time.sleep(RETRY_WAIT_SECONDS)
            else:
                raise RuntimeError(f"request failed after retries: {url} | {exc}")
    raise RuntimeError(f"request failed unexpectedly: {url} | {last_err}")


def stooq_last(symbol: str) -> Dict[str, float | str]:
    url = f"https://stooq.com/q/d/l/?s={symbol.lower()}.us&i=d"
    txt = fetch_text(url).strip().splitlines()
    rows = list(csv.DictReader(io.StringIO("\n".join(txt))))
    rows = [r for r in rows if r.get("Close")]
    if len(rows) < 2:
        raise RuntimeError(f"Not enough rows for Stooq symbol {symbol}")

    last = rows[-1]
    prev = rows[-6] if len(rows) >= 6 else rows[0]
    c = float(last["Close"])
    p = float(prev["Close"])
    chg5d = (c / p - 1.0) * 100 if p else 0.0
    return {"date": last["Date"], "chg5d": chg5d, "close": c}


def fred_last(series: str) -> Dict[str, float | str]:
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}"
    txt = fetch_text(url).strip().splitlines()
    rows = list(csv.DictReader(io.StringIO("\n".join(txt))))
    rows = [r for r in rows if r.get(series) and r[series] != "."]
    if not rows:
        raise RuntimeError(f"No rows for FRED series {series}")
    last = rows[-1]
    return {"date": last["observation_date"], "value": float(last[series])}


def parse_pub_date(value: str):
    if not value:
        return None
    try:
        dt = parsedate_to_datetime(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def clean_title_for_dedupe(title: str) -> str:
    title = title.lower()
    title = re.sub(r"\s+", " ", title)
    title = re.sub(r"[^\w\s]", "", title)
    return title.strip()


def normalize_source(source: str) -> str:
    s = (source or "").strip()
    if not s:
        return "Unknown"
    lower = s.lower()
    if "reuters" in lower:
        return "Reuters"
    if "financial times" in lower or lower == "ft":
        return "Financial Times"
    if "bloomberg" in lower:
        return "Bloomberg"
    if "wall street journal" in lower or lower == "wsj":
        return "WSJ"
    if "cnbc" in lower:
        return "CNBC"
    if "marketwatch" in lower:
        return "MarketWatch"
    if "yahoo" in lower:
        return "Yahoo"
    if "investmentnews" in lower:
        return "InvestmentNews"
    if "sec" in lower:
        return "SEC"
    return s


def title_matches_keywords(title: str, keywords: List[str]) -> bool:
    t = title.lower()
    return any(k.lower() in t for k in keywords)


def google_news_rss(query: str, max_items: int = 20) -> List[Dict[str, str]]:
    url = f"https://news.google.com/rss/search?q={quote_plus(query)}&hl=en-US&gl=US&ceid=US:en"
    xml_text = fetch_text(url)
    root = ET.fromstring(xml_text)

    results = []
    seen = set()

    for item in root.findall(".//item"):
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        pub_date = (item.findtext("pubDate") or "").strip()

        source = ""
        source_node = item.find("source")
        if source_node is not None and source_node.text:
            source = source_node.text.strip()

        dedupe_key = clean_title_for_dedupe(title)
        if not dedupe_key or dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        results.append({
            "title": title,
            "link": link,
            "date": pub_date,
            "source": normalize_source(source or "Google News RSS"),
        })
        if len(results) >= max_items:
            break
    return results


def filter_news_items(items: List[Dict[str, str]], rule: Dict[str, object]) -> List[Dict[str, str]]:
    now_utc = datetime.now(timezone.utc)
    lookback_days = int(rule["lookback_days"])
    cutoff = now_utc - timedelta(days=lookback_days)
    keywords = list(rule["must_have_any"])

    filtered = []
    seen_titles = set()

    for item in items:
        pub_dt = parse_pub_date(item.get("date", ""))
        if pub_dt is None or pub_dt < cutoff:
            continue

        title = item.get("title", "")
        if not title_matches_keywords(title, keywords):
            continue

        dedupe_key = clean_title_for_dedupe(title)
        if dedupe_key in seen_titles:
            continue
        seen_titles.add(dedupe_key)
        filtered.append(item)

    return filtered


def classify_news(rule: Dict[str, object], items: List[Dict[str, str]]) -> str:
    count = len(items)
    source_count = len({normalize_source(x.get("source", "")) for x in items if x.get("source")})

    if count >= int(rule["red_min_articles"]) and source_count >= int(rule["red_min_sources"]):
        return "红灯"
    if count >= int(rule["yellow_min_articles"]) and source_count >= int(rule["yellow_min_sources"]):
        return "黄灯"
    return "绿灯"


def format_news_rationale(label: str, items: List[Dict[str, str]], lookback_days: int) -> str:
    if not items:
        return f"最近 {lookback_days} 天未发现明确的{label}主流媒体高相关报道。"
    medias = sorted({normalize_source(x["source"]) for x in items if x.get("source")})
    return (
        f"最近 {lookback_days} 天共捕获 {len(items)} 条与{label}高度相关报道，"
        f"涉及 {len(medias)} 个来源；来源包括：{', '.join(medias[:5])}。"
        f"请结合 Evidence sheet 复核是否误报。"
    )


def ensure_workbook(master: Path, asset_template: Path) -> None:
    if master.exists():
        return
    if not asset_template.exists():
        raise FileNotFoundError(f"Template not found: {asset_template}")
    shutil.copy2(asset_template, master)


def ensure_sheet(wb: openpyxl.Workbook, title: str) -> Worksheet:
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title)


def build_row_map(ws: Worksheet) -> Dict[str, int]:
    result = {}
    for r in range(6, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if rid:
            result[str(rid)] = r
    return result


def ensure_checklist_columns(ws: Worksheet) -> Dict[str, int]:
    required = [
        "状态依据（具体数据+来源）",
        "证据链接 / Evidence IDs",
        "自动化方式",
        "主数据源",
        "备注 / 下一步动作",
    ]

    header_index = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(5, c).value
        if h:
            header_index[str(h)] = c

    next_col = ws.max_column
    for req in required:
        if req not in header_index:
            next_col += 1
            ws.cell(row=5, column=next_col, value=req)
            header_index[req] = next_col

    final_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(5, c).value
        if h:
            final_map[str(h)] = c
    return final_map


def reset_evidence_sheet(wb: openpyxl.Workbook) -> Worksheet:
    if "Evidence" in wb.sheetnames:
        wb.remove(wb["Evidence"])
    ws = wb.create_sheet("Evidence")
    headers = ["Evidence ID", "指标ID", "媒体", "标题", "日期", "链接"]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=h)
    return ws


def append_evidence(ws: Worksheet, items: List[Tuple[str, str, str, str, str, str]]) -> None:
    row = ws.max_row + 1
    for ev_id, rid, media, title, dt, link in items:
        ws.cell(row=row, column=1, value=ev_id)
        ws.cell(row=row, column=2, value=rid)
        ws.cell(row=row, column=3, value=media)
        ws.cell(row=row, column=4, value=title)
        ws.cell(row=row, column=5, value=dt)
        ws.cell(row=row, column=6, value=link)
        row += 1


def set_row(
    ws: Worksheet,
    row: int,
    colmap: Dict[str, int],
    status: str,
    rationale: str,
    evidence_ids: str,
    automation: str,
    source: str,
    note: str = "",
) -> None:
    ws.cell(row=row, column=8, value=status)
    ws.cell(row=row, column=9, value={"绿灯": 0, "黄灯": 1, "红灯": 2, "待更新": ""}.get(status, ""))
    ws.cell(row=row, column=colmap["状态依据（具体数据+来源）"], value=rationale)
    ws.cell(row=row, column=colmap["证据链接 / Evidence IDs"], value=evidence_ids)
    ws.cell(row=row, column=colmap["自动化方式"], value=automation)
    ws.cell(row=row, column=colmap["主数据源"], value=source)
    ws.cell(row=row, column=colmap["备注 / 下一步动作"], value=note)


def get_status(ws: Worksheet, row_map: Dict[str, int], rid: str) -> str:
    return str(ws.cell(row_map[rid], 8).value or "")


def classify_overall(checklist: Worksheet, row_map: Dict[str, int]) -> Tuple[str, str]:
    red_ids = [rid for rid in ROW_IDS if get_status(checklist, row_map, rid) == "红灯"]
    yellow_ids = [rid for rid in ROW_IDS if get_status(checklist, row_map, rid) == "黄灯"]

    ew_red = [rid for rid in red_ids if rid.startswith("EW-")]
    tr_red = [rid for rid in red_ids if rid.startswith("TR-")]
    sy_red = [rid for rid in red_ids if rid.startswith("SY-")]

    core_system_ids = {"TR-06", "SY-01", "SY-02", "SY-04"}
    core_red = [rid for rid in red_ids if rid in core_system_ids]

    # 红：系统性确认
    if ("SY-04" in red_ids and get_status(checklist, row_map, "TR-06") in {"黄灯", "红灯"}) or \
       ("TR-06" in red_ids and (get_status(checklist, row_map, "SY-01") == "红灯" or get_status(checklist, row_map, "SY-02") == "红灯")) or \
       (len([rid for rid in sy_red if rid in {"SY-01", "SY-02", "SY-04"}]) >= 2):
        return "红", LEVEL_STYLE["红"]["label"]

    # 橙：系统性风险临近
    if "TR-06" in red_ids or "SY-04" in red_ids or \
       (len(core_red) >= 1 and len(tr_red + sy_red) >= 2):
        return "橙", LEVEL_STYLE["橙"]["label"]

    # 黄：事件压力明显 / 传导开始
    if len(red_ids) >= 1 or len(yellow_ids) >= 4 or \
       len([rid for rid in tr_red if rid != "TR-06"]) >= 1 or \
       len(ew_red) >= 2:
        return "黄", LEVEL_STYLE["黄"]["label"]

    # 浅黄：预警升温
    if len(yellow_ids) >= 2 or len([rid for rid in yellow_ids if rid.startswith("EW-")]) >= 1:
        return "浅黄", LEVEL_STYLE["浅黄"]["label"]

    return "绿", LEVEL_STYLE["绿"]["label"]


def build_summary(
    overall_level: str,
    checklist: Worksheet,
    row_map: Dict[str, int],
    counts: Dict[str, int]
) -> str:
    red_ids = [rid for rid in ROW_IDS if get_status(checklist, row_map, rid) == "红灯"]
    yellow_ids = [rid for rid in ROW_IDS if get_status(checklist, row_map, rid) == "黄灯"]

    watch_ids = []
    for rid in ["TR-06", "SY-01", "SY-02", "SY-04"]:
        if get_status(checklist, row_map, rid) in {"黄灯", "红灯"}:
            watch_ids.append(rid)

    if overall_level == "绿":
        text = (
            f"当前私募信贷风险处于“绿（稳定）”阶段。16项监测中绿灯{counts['绿灯']}项，"
            f"黄灯{counts['黄灯']}项，红灯{counts['红灯']}项。"
            f"目前尚未看到明显的系统性危机迹象。"
        )
    elif overall_level == "浅黄":
        text = (
            f"当前私募信贷风险处于“浅黄（预警升温）”阶段。16项监测中绿灯{counts['绿灯']}项，"
            f"黄灯{counts['黄灯']}项，红灯{counts['红灯']}项。"
            f"异常开始增多，但仍主要停留在预警层，系统性确认尚不足。"
        )
    elif overall_level == "黄":
        text = (
            f"当前私募信贷风险处于“黄（压力上升）”阶段。16项监测中绿灯{counts['绿灯']}项，"
            f"黄灯{counts['黄灯']}项，红灯{counts['红灯']}项。"
            f"本轮主要压力来自事件与传导层，红灯项包括：{', '.join(red_ids[:3]) if red_ids else '无'}；"
            f"但系统确认层尚未形成红灯共振。后续重点观察：{', '.join(watch_ids[:4]) if watch_ids else 'TR-06, SY-01, SY-04'}。"
        )
    elif overall_level == "橙":
        text = (
            f"当前私募信贷风险处于“橙（系统风险临近）”阶段。16项监测中绿灯{counts['绿灯']}项，"
            f"黄灯{counts['黄灯']}项，红灯{counts['红灯']}项。"
            f"传导压力与系统确认压力已开始联动，核心关注：{', '.join(watch_ids[:4]) if watch_ids else 'TR-06, SY-01, SY-02, SY-04'}。"
        )
    else:
        text = (
            f"当前私募信贷风险处于“红（危机确认）”阶段。16项监测中绿灯{counts['绿灯']}项，"
            f"黄灯{counts['黄灯']}项，红灯{counts['红灯']}项。"
            f"系统层关键指标已形成共振，私募信贷风险进入危机阶段。"
        )

    return text[:290]


def write_dashboard(
    dashboard: Worksheet,
    counts: Dict[str, int],
    weighted_score: int,
    overall_level: str,
    overall_label: str,
    summary_text: str,
) -> None:
    # 左侧汇总
    dashboard["B5"] = datetime.now().strftime("%Y-%m-%d\n%H:%M:%S")
    dashboard["B5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dashboard["B6"] = 16
    dashboard["B7"] = counts["绿灯"]
    dashboard["B8"] = counts["黄灯"]
    dashboard["B9"] = counts["红灯"]
    dashboard["B10"] = counts["待更新"]
    dashboard["B11"] = weighted_score

    # 当前整体等级
    level_cell = dashboard["D6"]
    style = LEVEL_STYLE[overall_level]
    level_cell.value = f"{overall_level}\n{overall_label}"
    level_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    level_cell.font = Font(bold=True, size=24, color=style["font"])
    level_cell.fill = PatternFill("solid", fgColor=style["fill"])

    # 等级释义保持模板原文，不覆盖
    # 自动摘要
    dashboard["A21"] = summary_text
    dashboard["A21"].alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    parser = argparse.ArgumentParser(description="Update private credit monitor workbook using final dashboard layout")
    parser.add_argument("--workspace", default=".", help="Workspace directory")
    parser.add_argument("--master", default="private_credit_monitor_master_template.xlsx", help="Master workbook filename")
    args = parser.parse_args()

    workspace = Path(args.workspace).resolve()
    workspace.mkdir(parents=True, exist_ok=True)

    master = workspace / args.master
    latest = workspace / "private_credit_monitor_latest.xlsx"
    daily = workspace / f"private_credit_monitor_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    asset_template = Path(__file__).resolve().parents[1] / "assets" / TEMPLATE_NAME

    ensure_workbook(master, asset_template)

    wb = openpyxl.load_workbook(master)
    checklist = wb["Checklist"]
    dashboard = ensure_sheet(wb, "Dashboard")
    evidence_ws = reset_evidence_sheet(wb)

    colmap = ensure_checklist_columns(checklist)
    row_map = build_row_map(checklist)
    for rid in ROW_IDS:
        if rid not in row_map:
            raise RuntimeError(f"正式模板缺少指标行：{rid}")

    market = {}
    for symbol in ALL_STOOQ:
        try:
            market[symbol] = stooq_last(symbol)
        except Exception as exc:
            print(f"[WARN] Stooq failed for {symbol}: {exc}")

    fred = {}
    for alias, series in FRED_SERIES.items():
        try:
            fred[alias] = fred_last(series)
        except Exception as exc:
            print(f"[WARN] FRED failed for {series}: {exc}")

    evidence_rows = []
    evidence_counter = 1

    def add_news_metric(rid: str) -> None:
        nonlocal evidence_counter
        cfg = NEWS_RULES[rid]
        row = row_map[rid]
        try:
            raw_items = google_news_rss(str(cfg["query"]), max_items=20)
            items = filter_news_items(raw_items, cfg)
            status = classify_news(cfg, items)
            rationale = format_news_rationale(str(cfg["label"]), items, int(cfg["lookback_days"]))

            ev_ids = []
            for item in items[:5]:
                ev_id = f"EV-{evidence_counter:03d}"
                evidence_counter += 1
                ev_ids.append(ev_id)
                evidence_rows.append((ev_id, rid, normalize_source(item["source"]), item["title"], item["date"], item["link"]))

            set_row(checklist, row, colmap, status, rationale, ", ".join(ev_ids), str(cfg["automation"]), str(cfg["source"]))
        except Exception as exc:
            set_row(checklist, row, colmap, "待更新", f"新闻抓取失败：{exc}", "", str(cfg["automation"]), str(cfg["source"]))

    for rid in ["EW-01", "EW-05", "TR-01", "TR-02", "TR-04", "TR-05", "SY-03", "SY-05"]:
        add_news_metric(rid)

    # EW-02
    rid = "EW-02"
    row = row_map[rid]
    need = MANAGER_STOCKS + ["SPY", "XLF"]
    if all(x in market for x in need):
        mgr_avg = statistics.mean([market[s]["chg5d"] for s in MANAGER_STOCKS])
        bench_avg = statistics.mean([market["SPY"]["chg5d"], market["XLF"]["chg5d"]])
        rel = mgr_avg - bench_avg
        status = "绿灯" if rel > -1 else ("黄灯" if rel > -3 else "红灯")
        rationale = f"5日相对收益：资管股均值 {mgr_avg:.2f}% vs SPY/XLF 均值 {bench_avg:.2f}%，差值 {rel:.2f}%。来源：Stooq。"
        set_row(checklist, row, colmap, status, rationale, "", "数值直抓", "Stooq")
    else:
        missing = [x for x in need if x not in market]
        set_row(checklist, row, colmap, "待更新", f"Stooq 获取失败：{','.join(missing)}", "", "数值直抓", "Stooq")

    # EW-03
    rid = "EW-03"
    row = row_map[rid]
    if "BIZD" in market and "HYG" in market:
        rel_bdc = market["BIZD"]["chg5d"] - market["HYG"]["chg5d"]
        status = "绿灯" if rel_bdc >= -1 else ("黄灯" if rel_bdc >= -3 else "红灯")
        rationale = f"5日 BIZD {market['BIZD']['chg5d']:.2f}% vs HYG {market['HYG']['chg5d']:.2f}%，相对 {rel_bdc:.2f}%。作为 BDC 折价代理。"
        set_row(checklist, row, colmap, status, rationale, "", "代理数值", "Stooq")
    else:
        set_row(checklist, row, colmap, "待更新", "BDC/HY 代理数据获取失败", "", "代理数值", "Stooq")

    # EW-04
    rid = "EW-04"
    row = row_map[rid]
    if "HY_OAS" in fred and "BKLN" in market and "JBBB" in market:
        hy = fred["HY_OAS"]["value"]
        loan = market["BKLN"]["chg5d"]
        clo = market["JBBB"]["chg5d"]
        weak = sum([1 if hy > 4.0 else 0, 1 if loan < 0 else 0, 1 if clo < 0 else 0])
        status = "绿灯" if weak == 0 else ("黄灯" if weak == 1 else "红灯")
        rationale = f"HY OAS={hy:.2f}%({fred['HY_OAS']['date']}), BKLN 5日 {loan:.2f}%, JBBB 5日 {clo:.2f}%；弱项 {weak}/3。来源：FRED+Stooq。"
        set_row(checklist, row, colmap, status, rationale, "", "数值直抓", "FRED + Stooq")
    else:
        set_row(checklist, row, colmap, "待更新", "HY/Loan/CLO 代理数据不完整", "", "数值直抓", "FRED + Stooq")

    # TR-03
    rid = "TR-03"
    row = row_map[rid]
    if "BIZD" in market and "LQD" in market:
        rel = market["BIZD"]["chg5d"] - market["LQD"]["chg5d"]
        status = "绿灯" if rel > -1 else ("黄灯" if rel > -3 else "红灯")
        rationale = f"以 BDC 相对 IG 代理 NAV 压力：BIZD-LQD 5日差值 {rel:.2f}%。若持续弱于 IG，通常对应账面压力上升。"
        set_row(checklist, row, colmap, status, rationale, "", "代理数值", "Stooq")
    else:
        set_row(checklist, row, colmap, "待更新", "NAV 代理数据获取失败", "", "代理数值", "Stooq")

    # TR-06
    rid = "TR-06"
    row = row_map[rid]
    if "SOFR" in fred and "ANFCI" in fred:
        sofr = fred["SOFR"]["value"]
        anfci = fred["ANFCI"]["value"]
        status = "绿灯" if (sofr < 4 and anfci < 0.25) else ("黄灯" if (sofr < 5 and anfci < 0.75) else "红灯")
        rationale = f"SOFR={sofr:.3f}%({fred['SOFR']['date']}), ANFCI={anfci:.3f}({fred['ANFCI']['date']})。用短端资金与金融条件代理仓储融资/过桥成本。"
        set_row(checklist, row, colmap, status, rationale, "", "代理数值", "FRED")
    else:
        set_row(checklist, row, colmap, "待更新", "SOFR/ANFCI 获取失败", "", "代理数值", "FRED")

    # SY-01
    rid = "SY-01"
    row = row_map[rid]
    if "ANFCI" in fred and "KBE" in market:
        anfci = fred["ANFCI"]["value"]
        kbe = market["KBE"]["chg5d"]
        status = "绿灯" if anfci < 0 and kbe > -2 else ("黄灯" if anfci < 0.5 and kbe > -5 else "红灯")
        rationale = f"ANFCI={anfci:.3f}，KBE 5日={kbe:.2f}%。以金融条件和银行股表现代理授信收紧。"
        set_row(checklist, row, colmap, status, rationale, "", "代理数值", "FRED + Stooq")
    else:
        set_row(checklist, row, colmap, "待更新", "银行传导代理数据获取失败", "", "代理数值", "FRED + Stooq")

    # SY-02
    rid = "SY-02"
    row = row_map[rid]
    try:
        bank_news_raw = google_news_rss('"bank provisions" OR "loan loss provisions" OR "credit reserves" banks', max_items=12)
        bank_news = filter_news_items(bank_news_raw, {
            "lookback_days": 45,
            "must_have_any": ["provisions", "loan loss", "credit reserves", "reserve build", "reserve increase"],
            "yellow_min_articles": 2,
            "yellow_min_sources": 2,
            "red_min_articles": 4,
            "red_min_sources": 3,
        })
        bank_status = classify_news({
            "yellow_min_articles": 2,
            "yellow_min_sources": 2,
            "red_min_articles": 4,
            "red_min_sources": 3,
        }, bank_news)

        ev_ids = []
        for item in bank_news[:4]:
            ev_id = f"EV-{evidence_counter:03d}"
            evidence_counter += 1
            ev_ids.append(ev_id)
            evidence_rows.append((ev_id, rid, normalize_source(item["source"]), item["title"], item["date"], item["link"]))

        if "KBE" in market:
            kbe = market["KBE"]["chg5d"]
            if bank_status == "绿灯" and kbe > -2:
                status = "绿灯"
            elif bank_status == "红灯" or kbe < -5:
                status = "红灯"
            else:
                status = "黄灯"
            rationale = f"最近45天银行拨备相关新闻 {len(bank_news)} 条；KBE 5日 {kbe:.2f}%。结合新闻与银行股表现判断。"
        else:
            status = bank_status
            rationale = f"最近45天银行拨备相关新闻 {len(bank_news)} 条。缺少 KBE 价格时仅按新闻聚类判断。"

        set_row(checklist, row, colmap, status, rationale, ", ".join(ev_ids), "混合：新闻+代理数值", "Google News RSS + Stooq")
    except Exception as exc:
        set_row(checklist, row, colmap, "待更新", f"银行拨备监控失败：{exc}", "", "混合：新闻+代理数值", "Google News RSS + Stooq")

    # SY-04
    rid = "SY-04"
    row = row_map[rid]
    if "IG_OAS" in fred and "LQD" in market and "HYG" in market and "XLF" in market:
        ig = fred["IG_OAS"]["value"]
        lqd_hyg = market["LQD"]["chg5d"] - market["HYG"]["chg5d"]
        xlf = market["XLF"]["chg5d"]
        if ig < 1.4 and lqd_hyg >= 0 and xlf > -2:
            status = "绿灯"
        elif ig < 1.8 and xlf > -4:
            status = "黄灯"
        else:
            status = "红灯"
        rationale = f"IG OAS={ig:.2f}%({fred['IG_OAS']['date']}), LQD-HYG 5日相对={lqd_hyg:.2f}%，XLF 5日={xlf:.2f}%。"
        set_row(checklist, row, colmap, status, rationale, "", "数值直抓", "FRED + Stooq")
    else:
        set_row(checklist, row, colmap, "待更新", "IG/信用外溢数据不完整", "", "数值直抓", "FRED + Stooq")

    append_evidence(evidence_ws, evidence_rows)

    counts = {"绿灯": 0, "黄灯": 0, "红灯": 0, "待更新": 0}
    for rid in ROW_IDS:
        st = checklist.cell(row_map[rid], 8).value
        if st in counts:
            counts[st] += 1

    weighted_score = counts["黄灯"] * 1 + counts["红灯"] * 2
    overall_level, overall_label = classify_overall(checklist, row_map)
    summary_text = build_summary(overall_level, checklist, row_map, counts)

    write_dashboard(dashboard, counts, weighted_score, overall_level, overall_label, summary_text)

    wb.save(master)
    wb.save(latest)
    wb.save(daily)

    print(f"updated: {latest}")
    print(f"daily: {daily}")
    print(f"evidence_rows: {len(evidence_rows)}")
    print(f"status_counts: {counts}")
    print(f"overall_level: {overall_level}")
    print(f"overall_label: {overall_label}")


if __name__ == "__main__":
    main()
